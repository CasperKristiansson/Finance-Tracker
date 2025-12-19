"""Service layer for imports (stateless preview + atomic commit)."""

# pylint: disable=broad-exception-caught,too-many-lines

from __future__ import annotations

import base64
import io
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any, Dict, List, Optional, Tuple, cast
from uuid import UUID, uuid4

from openpyxl import load_workbook
from sqlalchemy import func, or_
from sqlmodel import Session, select

from ..models import (
    Account,
    Category,
    ImportRule,
    Subscription,
    TaxEvent,
    Transaction,
    TransactionImportBatch,
    TransactionLeg,
)
from ..schemas import (
    ImportCommitRequest,
    ImportCommitResponse,
    ImportPreviewRequest,
    ImportPreviewResponse,
)
from ..shared import (
    AccountType,
    BankImportType,
    CreatedSource,
    TaxEventType,
    TransactionType,
)
from .transaction import TransactionService


@dataclass
class CategorySuggestion:
    category_id: UUID | None
    category: Optional[str]
    confidence: float
    reason: Optional[str] = None


@dataclass
class SubscriptionSuggestion:
    subscription_id: UUID
    subscription_name: str
    confidence: float
    reason: Optional[str] = None


@dataclass
class RuleMatch:
    rule_id: UUID
    category_id: Optional[UUID]
    category_name: Optional[str]
    subscription_id: Optional[UUID]
    subscription_name: Optional[str]
    summary: str
    score: float
    rule_type: str


SUBSCRIPTION_SUGGESTION_THRESHOLD = 0.8


class ImportService:
    """Coordinates stateless import preview and atomic commit."""

    def __init__(self, session: Session):
        self.session = session

    def preview_import(self, payload: ImportPreviewRequest) -> dict[str, Any]:
        """Parse files and return draft rows + suggestions without persisting."""

        category_lookup_by_id = self._category_lookup_by_id()
        subscription_lookup_by_id = self._subscription_lookup_by_id()
        category_by_name = self._category_lookup()

        response_files: list[dict[str, Any]] = []
        response_rows: list[dict[str, Any]] = []
        rows_by_account: dict[UUID, list[tuple[UUID, str]]] = {}

        for file_payload in payload.files:
            file_id = uuid4()
            file_errors: list[tuple[int, str]] = []
            rows: list[dict[str, Any]] = []
            bank_import_type: BankImportType | None = None

            account = self.session.get(Account, file_payload.account_id)
            if account is None:
                file_errors.append((0, "Account not found"))
            else:
                bank_import_type = self._coerce_bank_import_type(account.bank_import_type)
                if bank_import_type is None:
                    file_errors.append((0, "Account has no bank import type configured"))
                else:
                    decoded = self._decode_base64(file_payload.content_base64)
                    rows, parse_errors = self._extract_bank_rows(
                        filename=file_payload.filename,
                        content=decoded,
                        bank_type=bank_import_type,
                    )
                    file_errors.extend(parse_errors)

            column_map: dict[str, str] | None = (
                {"date": "date", "description": "description", "amount": "amount"} if rows else None
            )
            file_errors.extend(self._validate_rows(rows, column_map))

            rule_matches = self._rule_matches(
                rows,
                column_map or {},
                category_lookup_by_id,
                subscription_lookup_by_id,
            )
            category_suggestions = self._suggest_rows(rows, column_map or {}, rule_matches)
            subscription_suggestions = self._suggest_subscriptions(
                rows, column_map or {}, rule_matches
            )
            transfers = self._match_transfers(rows, column_map or {})

            preview_rows = rows[:5]
            decorated_preview_rows: list[dict[str, Any]] = []
            for idx, row in enumerate(preview_rows):
                decorated = dict(row)
                suggestion = category_suggestions.get(idx)
                subscription_hint = subscription_suggestions.get(idx)
                transfer = transfers.get(idx)
                rule_match = rule_matches.get(idx)
                if suggestion:
                    decorated["suggested_category"] = suggestion.category
                    decorated["suggested_confidence"] = round(suggestion.confidence, 2)
                    if suggestion.reason:
                        decorated["suggested_reason"] = suggestion.reason
                if subscription_hint:
                    decorated["suggested_subscription_id"] = str(subscription_hint.subscription_id)
                    decorated["suggested_subscription_name"] = subscription_hint.subscription_name
                    decorated["suggested_subscription_confidence"] = round(
                        subscription_hint.confidence, 2
                    )
                    if subscription_hint.reason:
                        decorated["suggested_subscription_reason"] = subscription_hint.reason
                if transfer:
                    decorated["transfer_match"] = transfer
                if rule_match:
                    decorated["rule_applied"] = True
                    decorated["rule_type"] = rule_match.rule_type
                    decorated["rule_summary"] = rule_match.summary
                decorated_preview_rows.append(decorated)

            error_payloads = [{"row_number": row, "message": msg} for row, msg in file_errors]

            response_files.append(
                {
                    "id": file_id,
                    "filename": file_payload.filename,
                    "account_id": file_payload.account_id,
                    "bank_import_type": bank_import_type,
                    "row_count": len(rows),
                    "error_count": len(file_errors),
                    "errors": error_payloads,
                    "preview_rows": decorated_preview_rows,
                }
            )

            for idx, row in enumerate(rows, start=1):
                row_id = uuid4()
                occurred_at = str(row.get("date") or "")
                amount = str(row.get("amount") or "")
                description = str(row.get("description") or "")
                rows_by_account.setdefault(file_payload.account_id, []).append(
                    (row_id, description)
                )

                suggestion = category_suggestions.get(idx - 1)
                suggested_category_id: UUID | None = None
                suggested_category_name: str | None = None
                suggested_confidence: float | None = None
                suggested_reason: str | None = None
                if suggestion:
                    suggested_category_id = suggestion.category_id
                    suggested_category_name = suggestion.category
                    suggested_confidence = round(suggestion.confidence, 2)
                    suggested_reason = suggestion.reason
                    if suggested_category_id is None and suggestion.category:
                        category = category_by_name.get(suggestion.category.lower())
                        if category is not None:
                            suggested_category_id = category.id

                subscription_hint = subscription_suggestions.get(idx - 1)
                rule_match = rule_matches.get(idx - 1)

                response_rows.append(
                    {
                        "id": row_id,
                        "file_id": file_id,
                        "row_index": idx,
                        "account_id": file_payload.account_id,
                        "occurred_at": occurred_at,
                        "amount": amount,
                        "description": description,
                        "suggested_category_id": suggested_category_id,
                        "suggested_category_name": suggested_category_name,
                        "suggested_confidence": suggested_confidence,
                        "suggested_reason": suggested_reason,
                        "suggested_subscription_id": (
                            subscription_hint.subscription_id if subscription_hint else None
                        ),
                        "suggested_subscription_name": (
                            subscription_hint.subscription_name if subscription_hint else None
                        ),
                        "suggested_subscription_confidence": (
                            round(subscription_hint.confidence, 2) if subscription_hint else None
                        ),
                        "suggested_subscription_reason": (
                            subscription_hint.reason if subscription_hint else None
                        ),
                        "transfer_match": transfers.get(idx - 1),
                        "rule_applied": bool(rule_match),
                        "rule_type": rule_match.rule_type if rule_match else None,
                        "rule_summary": rule_match.summary if rule_match else None,
                    }
                )

        response_accounts: list[dict[str, Any]] = []
        for account_id, row_entries in rows_by_account.items():
            response_accounts.append(
                self._build_account_context(
                    account_id=account_id,
                    row_entries=row_entries,
                    category_lookup_by_id=category_lookup_by_id,
                )
            )

        return ImportPreviewResponse.model_validate(
            {"files": response_files, "rows": response_rows, "accounts": response_accounts}
        ).model_dump(mode="python")

    def _build_account_context(
        self,
        *,
        account_id: UUID,
        row_entries: list[tuple[UUID, str]],
        category_lookup_by_id: dict[UUID, Category],
        latest_limit: int = 40,
        similar_limit: int = 60,
        per_row_limit: int = 5,
    ) -> dict[str, Any]:
        recent = self._latest_transactions_for_account(account_id, limit=latest_limit)

        tokens: list[str] = []
        seen_tokens: set[str] = set()
        for _row_id, description in row_entries:
            for token in self._merchant_tokens(description):
                if token in seen_tokens:
                    continue
                seen_tokens.add(token)
                tokens.append(token)
                if len(tokens) >= 8:
                    break
            if len(tokens) >= 8:
                break

        similar: list[Transaction] = []
        if tokens:
            similar = self._similar_transactions_for_account(
                account_id,
                merchant_tokens=tokens,
                limit=similar_limit,
            )

        def tx_payload(tx: Transaction) -> dict[str, Any]:
            category = category_lookup_by_id.get(tx.category_id) if tx.category_id else None
            return {
                "id": tx.id,
                "account_id": account_id,
                "occurred_at": tx.occurred_at.isoformat(),
                "description": tx.description or "",
                "category_id": tx.category_id,
                "category_name": category.name if category else None,
            }

        recent_payloads = [tx_payload(tx) for tx in recent[:latest_limit]]
        similar_payloads = [tx_payload(tx) for tx in similar[:similar_limit]]

        similar_by_row: list[dict[str, Any]] = []
        similar_candidates = similar
        for row_id, description in row_entries:
            row_tokens = self._merchant_tokens(description)
            if not row_tokens:
                similar_by_row.append({"row_id": row_id, "transaction_ids": []})
                continue
            matches: list[UUID] = []
            for tx in similar_candidates:
                tx_desc = (tx.description or "").lower()
                if not tx_desc:
                    continue
                if any(token in tx_desc for token in row_tokens):
                    matches.append(tx.id)
                if len(matches) >= per_row_limit:
                    break
            similar_by_row.append({"row_id": row_id, "transaction_ids": matches})

        return {
            "account_id": account_id,
            "recent_transactions": recent_payloads,
            "similar_transactions": similar_payloads,
            "similar_by_row": similar_by_row,
        }

    def _latest_transactions_for_account(
        self, account_id: UUID, *, limit: int
    ) -> list[Transaction]:
        statement = (
            select(Transaction)
            .join(TransactionLeg, cast(Any, TransactionLeg.transaction_id) == Transaction.id)
            .where(cast(Any, TransactionLeg.account_id) == account_id)
            .where(cast(Any, Transaction.description).is_not(None))
            .order_by(cast(Any, Transaction.occurred_at).desc())
            .limit(limit)
        )
        return list(self.session.exec(statement).all())

    def _similar_transactions_for_account(
        self,
        account_id: UUID,
        *,
        merchant_tokens: list[str],
        limit: int,
    ) -> list[Transaction]:
        if not merchant_tokens:
            return []

        patterns = [f"%{token}%" for token in merchant_tokens]
        lowered = func.lower(cast(Any, Transaction.description))
        conditions = [lowered.like(pattern) for pattern in patterns]
        statement = (
            select(Transaction)
            .join(TransactionLeg, cast(Any, TransactionLeg.transaction_id) == Transaction.id)
            .where(cast(Any, TransactionLeg.account_id) == account_id)
            .where(cast(Any, Transaction.category_id).is_not(None))
            .where(cast(Any, Transaction.description).is_not(None))
            .where(or_(*conditions))
            .order_by(cast(Any, Transaction.occurred_at).desc())
            .limit(limit)
        )
        return list(self.session.exec(statement).all())

    def _merchant_tokens(self, description: str) -> list[str]:
        cleaned = re.sub(r"[^a-zA-Z0-9\\s]", " ", (description or "")).lower()
        parts = [part for part in cleaned.split() if len(part) >= 4]
        stop = {"card", "konto", "betalning", "payment", "purchase", "ab", "se", "sweden"}
        parts = [part for part in parts if part not in stop]
        return parts[:2]

    def commit_import(self, payload: ImportCommitRequest) -> dict[str, Any]:
        """Persist reviewed rows as transactions (all-or-nothing)."""

        if not payload.rows:
            raise ValueError("No rows provided")

        now = datetime.now(timezone.utc)
        batch = TransactionImportBatch(
            source_name=payload.note or "import",
            note=payload.note,
            created_at=now,
            updated_at=now,
        )
        self.session.add(batch)
        self.session.flush()

        transaction_service = TransactionService(self.session)
        offset_account = self._get_or_create_offset_account()

        created_ids: list[UUID] = []
        tax_events: list[TaxEvent] = []

        for row in payload.rows:
            if row.delete:
                continue

            occurred_at = self._parse_date(row.occurred_at)
            if occurred_at is None:
                raise ValueError("Date is not a valid ISO date")

            if not self._is_decimal(row.amount):
                raise ValueError("Amount must be numeric")

            amount = Decimal(str(row.amount))
            description = str(row.description or "").strip()
            if not description:
                raise ValueError("Description is required")

            tax_event_type: TaxEventType | None = row.tax_event_type
            category_id = row.category_id if tax_event_type is None else None
            subscription_id = row.subscription_id if tax_event_type is None else None

            target_account_id = row.account_id
            counterparty_account_id = row.transfer_account_id

            abs_amount = abs(amount)
            if tax_event_type is not None:
                cash_delta = abs_amount if tax_event_type == TaxEventType.REFUND else -abs_amount
                legs = [
                    TransactionLeg(account_id=target_account_id, amount=cash_delta),
                    TransactionLeg(account_id=offset_account.id, amount=-cash_delta),
                ]
                counterparty_account_id = None
            elif counterparty_account_id is not None:
                legs = [
                    TransactionLeg(account_id=target_account_id, amount=amount),
                    TransactionLeg(account_id=counterparty_account_id, amount=-amount),
                ]
            else:
                legs = [
                    TransactionLeg(account_id=target_account_id, amount=amount),
                    TransactionLeg(account_id=offset_account.id, amount=-amount),
                ]

            transaction = Transaction(
                category_id=category_id,
                transaction_type=TransactionType.TRANSFER,
                description=description,
                notes=None,
                external_id=None,
                occurred_at=occurred_at,
                posted_at=occurred_at,
                subscription_id=subscription_id,
                created_source=CreatedSource.IMPORT,
                import_batch_id=batch.id,
                created_at=now,
                updated_at=now,
            )

            created_tx = transaction_service.create_transaction(
                transaction,
                legs,
                import_batch=batch,
                commit=False,
            )
            if created_tx.id is None:  # pragma: no cover - defensive
                raise ValueError("Transaction was not persisted")
            created_ids.append(created_tx.id)

            if tax_event_type is not None:
                tax_events.append(
                    TaxEvent(
                        transaction_id=created_tx.id,
                        event_type=tax_event_type,
                    )
                )
            else:
                self._record_rule_from_row(
                    description,
                    amount,
                    occurred_at,
                    category_id,
                    subscription_id,
                )

        if tax_events:
            self.session.add_all(tax_events)

        return ImportCommitResponse(
            import_batch_id=batch.id,
            transaction_ids=created_ids,
        ).model_dump(mode="python")

    def _decode_base64(self, payload: str) -> bytes:
        try:
            return base64.b64decode(payload, validate=True)
        except Exception as exc:
            raise ValueError("Unable to decode file content") from exc

    def _coerce_bank_import_type(self, value: Any) -> BankImportType | None:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        try:
            return BankImportType(text)
        except ValueError:
            return None

    def _extract_bank_rows(
        self, *, filename: str, content: bytes, bank_type: BankImportType
    ) -> tuple[List[dict[str, Any]], List[Tuple[int, str]]]:
        name = filename.lower()
        if not name.endswith(".xlsx"):
            return ([], [(0, "Unsupported file type; only XLSX exports are accepted")])

        try:
            workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            return ([], [(0, f"Unable to read XLSX file: {exc}")])

        sheet = workbook.active
        if sheet is None:
            return ([], [(0, "XLSX workbook has no active sheet")])

        if bank_type == BankImportType.CIRCLE_K_MASTERCARD:
            return self._parse_circle_k_mastercard(sheet)
        if bank_type == BankImportType.SEB:
            return self._parse_seb(sheet)
        if bank_type == BankImportType.SWEDBANK:
            return self._parse_swedbank(sheet)
        return ([], [(0, "Unknown bank type")])

    def _parse_circle_k_mastercard(
        self, sheet
    ) -> tuple[List[dict[str, Any]], List[Tuple[int, str]]]:
        header_map: Dict[str, int] | None = None
        rows: List[dict[str, Any]] = []
        errors: List[Tuple[int, str]] = []

        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            cleaned = [self._clean_value(cell) for cell in row]
            if header_map is None:
                candidate_headers = {
                    self._clean_header(val): pos for pos, val in enumerate(row) if val is not None
                }
                if {"datum", "belopp"}.issubset(candidate_headers.keys()):
                    header_map = candidate_headers
                continue
            if not any(cleaned):
                continue
            first_cell = cleaned[0].lower() if cleaned and isinstance(cleaned[0], str) else ""
            if first_cell == "datum":
                candidate_headers = {
                    self._clean_header(val): pos for pos, val in enumerate(row) if val is not None
                }
                if {"datum", "belopp"}.issubset(candidate_headers.keys()):
                    header_map = candidate_headers
                continue
            if "totalt belopp" in first_cell:
                continue
            if "summa" in first_cell or (
                len(cleaned) > 2 and isinstance(cleaned[2], str) and "summa" in cleaned[2].lower()
            ):
                continue

            try:
                date_idx = header_map.get("datum", 0)
                date_text = cleaned[date_idx] if date_idx < len(cleaned) else ""

                description_idx = header_map.get("specifikation", 2)
                description = cleaned[description_idx] if description_idx < len(cleaned) else ""

                location_idx = header_map.get("ort", 3)
                location = cleaned[location_idx] if location_idx < len(cleaned) else ""
                if location:
                    description = f"{description} ({location})" if description else str(location)

                amount_idx = header_map.get("belopp", 6)
                amount_raw = cleaned[amount_idx] if amount_idx < len(cleaned) else ""

                occurred_at = self._parse_date(str(date_text))
                if occurred_at is None:
                    if date_text and description and amount_raw != "":
                        raise ValueError("invalid date")
                    continue
                if amount_raw == "":
                    continue
                amount = self._parse_decimal_value(amount_raw)
                if amount is None:
                    raise ValueError("invalid amount")
                amount = -abs(amount)
                rows.append(
                    {
                        "date": occurred_at.isoformat(),
                        "description": description,
                        "amount": str(amount),
                    }
                )
            except Exception as exc:
                errors.append((idx, f"Unable to parse row: {exc}"))

        if header_map is None:
            errors.append((0, "Circle K Mastercard export is missing the expected headers"))

        return rows, errors

    def _parse_seb(self, sheet) -> tuple[List[dict[str, Any]], List[Tuple[int, str]]]:
        header_index = None
        headers: Dict[str, int] = {}
        rows: List[dict[str, Any]] = []
        errors: List[Tuple[int, str]] = []

        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            header_map = {self._clean_header(val): pos for pos, val in enumerate(row) if val}
            if not header_index and {"bokförd", "insättningar/uttag"}.issubset(header_map.keys()):
                header_index = idx
                headers = header_map
                continue
            if header_index is None or idx <= header_index:
                continue
            cleaned = [self._clean_value(cell) for cell in row]
            if not any(cleaned):
                continue
            try:
                date_text = cleaned[headers.get("bokförd", 0)] if headers else cleaned[0]
                if not date_text:
                    continue
                occurred_at = self._parse_date(str(date_text))
                if occurred_at is None:
                    raise ValueError("invalid date")
                description = cleaned[headers.get("text", 2)] if headers else ""
                if not description:
                    description = cleaned[headers.get("typ", 3)] if headers else ""
                amount_raw = cleaned[headers.get("insättningar/uttag", 5)] if headers else ""
                amount = self._parse_decimal_value(amount_raw)
                if amount is None:
                    raise ValueError("invalid amount")
                rows.append(
                    {
                        "date": occurred_at.isoformat(),
                        "description": str(description),
                        "amount": str(amount),
                    }
                )
            except Exception as exc:
                errors.append((idx, f"Unable to parse row: {exc}"))

        if header_index is None:
            errors.append((0, "SEB export is missing the expected headers"))

        return rows, errors

    def _parse_swedbank(self, sheet) -> tuple[List[dict[str, Any]], List[Tuple[int, str]]]:
        header_index = None
        headers: Dict[str, int] = {}
        rows: List[dict[str, Any]] = []
        errors: List[Tuple[int, str]] = []

        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            header_map = {self._clean_header(val): pos for pos, val in enumerate(row) if val}
            if not header_index and {"bokföringsdag", "belopp"}.issubset(header_map.keys()):
                header_index = idx
                headers = header_map
                continue
            if header_index is None or idx <= header_index:
                continue
            cleaned = [self._clean_value(cell) for cell in row]
            if not any(cleaned):
                continue
            try:
                date_text = cleaned[headers.get("bokföringsdag", 1)]
                occurred_at = self._parse_date(str(date_text))
                if occurred_at is None:
                    raise ValueError("invalid date")
                ref = cleaned[headers.get("referens", 4)] if headers else ""
                desc = cleaned[headers.get("beskrivning", 5)] if headers else ""
                description = f"{ref} {desc}".strip() or desc or ref
                amount_raw = cleaned[headers.get("belopp", 6)] if headers else ""
                amount = self._parse_decimal_value(amount_raw)
                if amount is None:
                    raise ValueError("invalid amount")
                rows.append(
                    {
                        "date": occurred_at.isoformat(),
                        "description": description,
                        "amount": str(amount),
                    }
                )
            except Exception as exc:
                errors.append((idx, f"Unable to parse row: {exc}"))

        if header_index is None:
            errors.append((0, "Swedbank export is missing the expected headers"))

        return rows, errors

    def _validate_rows(
        self, rows: List[dict[str, Any]], column_map: Optional[Dict[str, str]]
    ) -> List[Tuple[int, str]]:
        if not rows:
            return []

        errors: List[Tuple[int, str]] = []

        if column_map is None:
            errors.append((0, "Missing required columns: date, description, amount"))
            return errors

        missing_columns = [field for field, header in column_map.items() if header is None]
        if missing_columns:
            errors.append((0, f"Missing required columns: {', '.join(missing_columns)}"))
            return errors

        for idx, row in enumerate(rows, start=1):
            missing_fields = [
                field for field, header in column_map.items() if not row.get(header, "")
            ]
            if missing_fields:
                errors.append((idx, f"Missing required fields: {', '.join(missing_fields)}"))
                continue

            amount_value = row[column_map["amount"]]
            if not self._is_decimal(amount_value):
                errors.append((idx, "Amount must be numeric"))

            date_value = row[column_map["date"]]
            if not self._is_date_like(date_value):
                errors.append((idx, "Date is not a valid ISO date"))

        return errors

    def _rule_matches(
        self,
        rows: List[dict],
        column_map: Dict[str, str],
        category_lookup: Dict[UUID, Category],
        subscription_lookup: Dict[UUID, Subscription],
    ) -> Dict[int, RuleMatch]:
        rules = self._active_rules()
        description_header = column_map.get("description")
        if not rules or not description_header:
            return {}

        amount_header = column_map.get("amount")
        date_header = column_map.get("date")
        matches: Dict[int, RuleMatch] = {}
        for idx, row in enumerate(rows):
            description = str(row.get(description_header, "") or "")
            amount = self._safe_decimal(row.get(amount_header)) if amount_header else None
            occurred_at = self._parse_date(str(row.get(date_header, ""))) if date_header else None
            best: RuleMatch | None = None
            for rule in rules:
                candidate = self._score_rule(
                    rule,
                    description,
                    amount,
                    occurred_at,
                    category_lookup,
                    subscription_lookup,
                )
                if candidate and (best is None or candidate.score > best.score):
                    best = candidate
            if best:
                matches[idx] = best
        return matches

    # pylint: disable=too-many-positional-arguments
    def _score_rule(
        self,
        rule: ImportRule,
        description: str,
        amount: Optional[Decimal],
        occurred_at: Optional[datetime],
        category_lookup: Dict[UUID, Category],
        subscription_lookup: Dict[UUID, Subscription],
    ) -> Optional[RuleMatch]:
        if not rule.is_active:
            return None

        normalized = description.lower()
        matcher = rule.matcher_text.lower()
        if matcher not in normalized:
            return None

        score = 0.7
        summary: List[str] = [f"matched '{rule.matcher_text}'"]

        if rule.matcher_day_of_month and occurred_at:
            if abs(occurred_at.day - rule.matcher_day_of_month) <= 1:
                score += 0.15
                summary.append(f"day≈{rule.matcher_day_of_month}")
            else:
                return None

        if rule.matcher_amount is not None and amount is not None:
            target = abs(rule.matcher_amount)
            delta = (abs(amount) - target).copy_abs()
            tolerance = rule.amount_tolerance or Decimal("0")
            if delta <= tolerance:
                score += 0.15
                summary.append(f"amount within ±{tolerance}")
            else:
                return None

        category = category_lookup.get(rule.category_id) if rule.category_id else None
        subscription = (
            subscription_lookup.get(rule.subscription_id) if rule.subscription_id else None
        )
        if category is None and subscription is None:
            return None

        rule_type = "category"
        if category and subscription:
            rule_type = "category+subscription"
        elif subscription and not category:
            rule_type = "subscription"

        return RuleMatch(
            rule_id=rule.id,
            category_id=category.id if category else None,
            category_name=category.name if category else None,
            subscription_id=subscription.id if subscription else None,
            subscription_name=subscription.name if subscription else None,
            summary="; ".join(summary),
            score=float(score),
            rule_type=rule_type,
        )

    def _active_rules(self) -> List[ImportRule]:
        statement = select(ImportRule).where(cast(Any, ImportRule.is_active).is_(True))
        return list(self.session.exec(statement).all())

    def _category_lookup_by_id(self) -> dict[UUID, Category]:
        categories = self.session.exec(select(Category)).all()
        return {cat.id: cat for cat in categories if getattr(cat, "id", None) is not None}

    def _subscription_lookup_by_id(self) -> dict[UUID, Subscription]:
        subscriptions = self.session.exec(select(Subscription)).all()
        return {sub.id: sub for sub in subscriptions if getattr(sub, "id", None) is not None}

    def _suggest_rows(
        self,
        rows: List[dict],
        column_map: Dict[str, str],
        rule_matches: Optional[Dict[int, RuleMatch]] = None,
    ) -> Dict[int, CategorySuggestion]:
        if not column_map or not column_map.get("description") or not column_map.get("amount"):
            return {}
        heuristic: Dict[int, CategorySuggestion] = {}
        locked: set[int] = set()
        if rule_matches:
            for idx, match in rule_matches.items():
                if match.category_name:
                    heuristic[idx] = CategorySuggestion(
                        category_id=match.category_id,
                        category=match.category_name,
                        confidence=0.95,
                        reason=match.summary or "Rule match",
                    )
                    locked.add(idx)

        for idx, row in enumerate(rows):
            if idx in locked:
                continue
            heuristic[idx] = self._suggest_category_heuristic(row, column_map)

        return heuristic

    def _suggest_category_heuristic(
        self, row: dict, column_map: Dict[str, str]
    ) -> CategorySuggestion:
        description = str(row.get(column_map["description"], ""))
        amount_text = str(row.get(column_map["amount"], ""))

        normalized_desc = description.lower()
        keyword_map = {
            "rent": "Rent",
            "salary": "Salary",
            "payroll": "Salary",
            "grocery": "Groceries",
            "market": "Groceries",
            "uber": "Transport",
            "lyft": "Transport",
            "electric": "Utilities",
            "water": "Utilities",
            "internet": "Utilities",
        }
        for keyword, category in keyword_map.items():
            if keyword in normalized_desc:
                return CategorySuggestion(category_id=None, category=category, confidence=0.65)

        return CategorySuggestion(
            category_id=None,
            category=None,
            confidence=0.3,
            reason=f"No signal for {amount_text}",
        )

    def _suggest_subscriptions(
        self,
        rows: List[dict],
        column_map: Dict[str, str],
        rule_matches: Optional[Dict[int, RuleMatch]] = None,
    ) -> Dict[int, SubscriptionSuggestion]:
        if not rows:
            return {}
        description_header = column_map.get("description")
        amount_header = column_map.get("amount")
        date_header = column_map.get("date")
        if not description_header:
            return {}

        subscriptions = self._active_subscriptions()
        if not subscriptions:
            return {}
        last_amounts = self._subscription_amount_lookup()

        suggestions: Dict[int, SubscriptionSuggestion] = {}
        locked: set[int] = set()
        if rule_matches:
            for idx, match in rule_matches.items():
                if match.subscription_id:
                    suggestions[idx] = SubscriptionSuggestion(
                        subscription_id=match.subscription_id,
                        subscription_name=match.subscription_name or "Matched rule",
                        confidence=0.99,
                        reason=match.summary or "Rule match",
                    )
                    locked.add(idx)
        for idx, row in enumerate(rows):
            if idx in locked:
                continue
            description = str(row.get(description_header, "") or "")
            amount = self._safe_decimal(row.get(amount_header)) if amount_header else None
            occurred_at = self._parse_date(str(row.get(date_header, ""))) if date_header else None

            best: SubscriptionSuggestion | None = None
            for subscription in subscriptions:
                candidate = self._score_subscription(
                    subscription,
                    description,
                    amount,
                    occurred_at,
                    last_amounts.get(subscription.id),
                )
                if candidate is None:
                    continue
                if best is None or candidate.confidence > best.confidence:
                    best = candidate
            if best and best.confidence >= SUBSCRIPTION_SUGGESTION_THRESHOLD:
                suggestions[idx] = best
        return suggestions

    def _score_subscription(  # pylint: disable=too-many-positional-arguments
        self,
        subscription: Subscription,
        description: str,
        amount: Optional[Decimal],
        occurred_at: Optional[datetime],
        last_amount: Optional[Decimal],
    ) -> Optional[SubscriptionSuggestion]:
        text_match, text_reason = self._match_subscription_text(
            subscription.matcher_text, description
        )
        if not text_match:
            return None

        confidence = 0.82 if text_reason == "regex" else 0.8
        reasons = [f"{text_reason} match"]

        if subscription.matcher_day_of_month and occurred_at:
            if occurred_at.day == subscription.matcher_day_of_month:
                confidence += 0.1
                reasons.append("day-of-month aligns")
            else:
                confidence -= 0.05

        if subscription.matcher_amount_tolerance is not None and amount is not None:
            if last_amount is not None:
                delta = (abs(amount) - abs(last_amount)).copy_abs()
                if delta <= subscription.matcher_amount_tolerance:
                    confidence += 0.08
                    reasons.append("amount within tolerance of last charge")
                else:
                    return None
            else:
                reasons.append("tolerance provided but no history; skipping amount check")

        confidence = min(confidence, 0.98)

        return SubscriptionSuggestion(
            subscription_id=subscription.id,
            subscription_name=subscription.name,
            confidence=confidence,
            reason="; ".join(reasons),
        )

    def _match_subscription_text(self, pattern: str, description: str) -> tuple[bool, str]:
        normalized = description.lower()
        try:
            if re.search(pattern, description, flags=re.IGNORECASE):
                return True, "regex"
        except re.error:
            pass
        if pattern.lower() in normalized:
            return True, "substring"
        return False, ""

    def _active_subscriptions(self) -> List[Subscription]:
        statement = select(Subscription).where(cast(Any, Subscription.is_active).is_(True))
        return list(self.session.exec(statement).all())

    def _subscription_amount_lookup(self) -> dict[UUID, Decimal]:
        latest = (
            select(
                cast(Any, Transaction.id).label("txn_id"),
                cast(Any, Transaction.subscription_id).label("subscription_id"),
                func.row_number()
                .over(
                    partition_by=cast(Any, Transaction.subscription_id),
                    order_by=cast(Any, Transaction.occurred_at).desc(),
                )
                .label("rn"),
            ).where(cast(Any, Transaction.subscription_id).isnot(None))
        ).subquery()

        amounts = (
            select(
                latest.c.subscription_id,
                func.max(func.abs(TransactionLeg.amount)).label("amount"),
            )
            .join(TransactionLeg, cast(Any, TransactionLeg.transaction_id == latest.c.txn_id))
            .where(latest.c.rn == 1)
            .group_by(latest.c.subscription_id)
        )
        results = self.session.exec(amounts).all()
        lookup: dict[UUID, Decimal] = {}
        for sub_id, amount in results:
            if sub_id is None or amount is None:
                continue
            lookup[sub_id] = Decimal(str(amount))
        return lookup

    def _match_transfers(
        self, rows: List[dict[str, Any]], column_map: Dict[str, str]
    ) -> Dict[int, dict[str, Any]]:
        if not rows:
            return {}

        amount_header = column_map.get("amount")
        date_header = column_map.get("date")
        if not amount_header:
            return {}

        entries: List[Dict[str, Any]] = []
        for idx, row in enumerate(rows):
            try:
                amount = Decimal(str(row.get(amount_header, "")))
            except Exception:
                continue
            date_value = self._parse_date(str(row.get(date_header, ""))) if date_header else None
            entries.append({"idx": idx, "amount": amount, "date": date_value})

        matches: Dict[int, dict[str, Any]] = {}
        used: set[int] = set()
        for entry in entries:
            if entry["idx"] in used:
                continue
            for other in entries:
                if other["idx"] in used or other["idx"] == entry["idx"]:
                    continue
                if (entry["amount"] + other["amount"]) != 0:
                    continue
                if entry["date"] and other["date"]:
                    delta = abs((entry["date"] - other["date"]).days)
                    if delta > 2:
                        continue
                match_payload = {
                    "paired_with": other["idx"] + 1,
                    "reason": "Matched transfer by amount and date proximity",
                }
                matches[entry["idx"]] = match_payload
                matches[other["idx"]] = {
                    "paired_with": entry["idx"] + 1,
                    "reason": "Matched transfer by amount and date proximity",
                }
                used.update({entry["idx"], other["idx"]})
                break
        return matches

    def _get_or_create_offset_account(self) -> Account:
        if hasattr(self, "_offset_account"):
            return getattr(self, "_offset_account")

        statement = select(Account).where(
            cast(Any, Account.is_active).is_(False), Account.name == "Offset"
        )
        account = self.session.exec(statement).one_or_none()
        if account is None:
            account = Account(
                name="Offset",
                account_type=AccountType.NORMAL,
                is_active=False,
            )
            self.session.add(account)
            self.session.flush()
        else:
            account.name = account.name or "Offset"
        setattr(self, "_offset_account", account)
        return account

    def _category_lookup(self) -> dict[str, Category]:
        statement = select(Category)
        categories = self.session.exec(statement).all()
        return {cat.name.lower(): cat for cat in categories}

    # pylint: disable=too-many-positional-arguments
    def _record_rule_from_row(
        self,
        description: Optional[str],
        amount: Optional[Decimal],
        occurred_at: Optional[datetime],
        category_id: Optional[UUID],
        subscription_id: Optional[UUID],
    ) -> None:
        if not description or (category_id is None and subscription_id is None):
            return

        matcher_text = description.lower().strip()
        if not matcher_text:
            return

        amount_abs = abs(amount) if amount is not None else None
        tolerance = self._derive_amount_tolerance(amount_abs)
        day_of_month = occurred_at.day if occurred_at else None

        existing = self.session.exec(
            select(ImportRule).where(func.lower(ImportRule.matcher_text) == matcher_text)
        ).one_or_none()

        if existing:
            if category_id:
                existing.category_id = category_id
            if subscription_id:
                existing.subscription_id = subscription_id
            if amount_abs is not None and existing.matcher_amount is None:
                existing.matcher_amount = amount_abs
            if tolerance and existing.amount_tolerance is None:
                existing.amount_tolerance = tolerance
            if day_of_month and existing.matcher_day_of_month is None:
                existing.matcher_day_of_month = day_of_month
            existing.updated_at = datetime.now(timezone.utc)
            return

        rule = ImportRule(
            matcher_text=matcher_text,
            matcher_amount=amount_abs,
            amount_tolerance=tolerance,
            matcher_day_of_month=day_of_month,
            category_id=category_id,
            subscription_id=subscription_id,
        )
        self.session.add(rule)

    def _derive_amount_tolerance(self, amount: Optional[Decimal]) -> Optional[Decimal]:
        if amount is None:
            return None
        amount_abs = abs(amount)
        baseline = Decimal("1.00")
        percent = (amount_abs * Decimal("0.02")).quantize(Decimal("0.01"))
        return max(baseline, percent)

    def _parse_date(self, value: str) -> Optional[datetime]:
        if not value:
            return None
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            return None

    def _clean_header(self, header: object) -> str:
        return str(header).strip().lower().replace(" ", "_") if header is not None else ""

    def _clean_value(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, (int, float, Decimal)):
            return str(value)
        return str(value).strip()

    def _parse_decimal_value(self, value: object) -> Optional[Decimal]:
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        text = str(value).strip()
        if not text:
            return None
        cleaned = text.replace("\u2212", "-").replace("−", "-").replace("\xa0", "").replace(" ", "")
        match = re.search(r"-?[\d.,]+", cleaned)
        if not match:
            return None
        numeric = match.group(0)
        if "," in numeric and "." in numeric:
            if numeric.rfind(",") > numeric.rfind("."):
                numeric = numeric.replace(".", "")
                numeric = numeric.replace(",", ".")
            else:
                numeric = numeric.replace(",", "")
        else:
            numeric = numeric.replace(",", ".")
        try:
            return Decimal(numeric)
        except Exception:
            return None

    def _is_decimal(self, value: str) -> bool:
        try:
            Decimal(str(value))
        except Exception:
            return False
        return True

    def _safe_decimal(self, value: Any) -> Optional[Decimal]:
        try:
            return Decimal(str(value))
        except Exception:
            return None

    def _is_date_like(self, value: str) -> bool:
        if not value:
            return False
        text = str(value)
        try:
            datetime.fromisoformat(text.replace("Z", "+00:00"))
            return True
        except ValueError:
            return False


__all__ = ["ImportService", "CategorySuggestion", "SubscriptionSuggestion", "RuleMatch"]
