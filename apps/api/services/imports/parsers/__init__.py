"""Bank-specific parser strategies for import previews."""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Callable, Dict, Optional, Protocol

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ....shared import BankImportType
from ..utils import parse_iso_date
from .circle_k_mastercard import parse_circle_k_mastercard
from .seb import parse_seb
from .swedbank import parse_swedbank
from .types import NormalizedRow, ParseResult


class BankParser(Protocol):
    def __call__(
        self, sheet: Any, *, parse_date: Callable[[str], Optional[datetime]]
    ) -> ParseResult: ...


_PARSER_MAP: Dict[BankImportType, BankParser] = {
    BankImportType.CIRCLE_K_MASTERCARD: parse_circle_k_mastercard,
    BankImportType.SEB: parse_seb,
    BankImportType.SWEDBANK: parse_swedbank,
}


def parse_bank_rows(*, filename: str, content: bytes, bank_type: BankImportType) -> ParseResult:
    """Load the workbook and delegate to the appropriate bank parser."""

    name = filename.lower()
    if not name.endswith(".xlsx"):
        return ([], [(0, "Unsupported file type; only XLSX exports are accepted")])

    parser = _PARSER_MAP.get(bank_type)
    if parser is None:
        return ([], [(0, "Unknown bank type")])

    try:
        workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError) as exc:  # pragma: no cover - defensive
        return ([], [(0, f"Unable to read XLSX file: {exc}")])

    sheet = workbook.active
    if sheet is None:
        return ([], [(0, "XLSX workbook has no active sheet")])

    rows, errors = parser(sheet, parse_date=parse_iso_date)
    if len(rows) > 1:
        return rows, errors

    try:
        fallback = load_workbook(filename=io.BytesIO(content), read_only=False, data_only=True)
    except (InvalidFileException, OSError, ValueError):  # pragma: no cover - defensive
        return rows, errors

    fallback_sheet = fallback.active
    if fallback_sheet is None:
        return rows, errors

    fallback_rows, fallback_errors = parser(fallback_sheet, parse_date=parse_iso_date)
    if len(fallback_rows) > len(rows):
        return fallback_rows, fallback_errors
    return rows, errors


__all__ = ["NormalizedRow", "ParseResult", "parse_bank_rows"]
