from __future__ import annotations

import base64
import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterator
from uuid import UUID

import pytest
from openpyxl import load_workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel

from apps.api.handlers.imports import preview_imports, reset_handler_state
from apps.api.models import Account
from apps.api.shared import (
    AccountType,
    configure_engine,
    get_default_user_id,
    get_engine,
    scope_session_to_user,
)


@pytest.fixture(autouse=True)
def configure_sqlite(monkeypatch: pytest.MonkeyPatch) -> Iterator[None]:
    monkeypatch.setenv("DATABASE_URL", "sqlite://")
    reset_handler_state()
    configure_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    engine = get_engine()
    SQLModel.metadata.create_all(engine)
    try:
        yield
    finally:
        SQLModel.metadata.drop_all(engine)


def _create_account(*, bank_import_type: str) -> UUID:
    engine = get_engine()
    with Session(engine) as session:
        scope_session_to_user(session, get_default_user_id())
        account = Account(
            name=f"Account ({bank_import_type})",
            account_type=AccountType.NORMAL,
            bank_import_type=bank_import_type,
            is_active=True,
        )
        session.add(account)
        session.commit()
        session.refresh(account)
        return account.id


def _b64_file(path: Path) -> str:
    return base64.b64encode(path.read_bytes()).decode()


def _json_body(response: dict) -> dict:
    return json.loads(response["body"])


def _coerce_iso_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if len(text) == 10 and text[4] == "-" and text[7] == "-":
        return text
    return None


def _find_header_row(*, sheet, required: set[str]) -> tuple[int | None, dict[str, int]]:
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        headers = {
            str(val).strip().lower(): pos
            for pos, val in enumerate(row)
            if val is not None and str(val).strip()
        }
        if required.issubset(headers.keys()):
            return idx, headers
    return None, {}


def _expected_transaction_count_and_sum(
    *, bank_import_type: str, file_path: Path
) -> tuple[int, Decimal]:
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        return 0, Decimal("0")

    if bank_import_type == "swedbank":
        header_idx, headers = _find_header_row(sheet=sheet, required={"bokföringsdag", "belopp"})
        if header_idx is None:
            return 0, Decimal("0")
        date_idx = headers["bokföringsdag"]
        amount_idx = headers["belopp"]
    elif bank_import_type == "seb":
        header_idx, headers = _find_header_row(
            sheet=sheet, required={"bokförd", "insättningar/uttag"}
        )
        if header_idx is None:
            return 0, Decimal("0")
        date_idx = headers["bokförd"]
        amount_idx = headers["insättningar/uttag"]
    elif bank_import_type == "circle_k_mastercard":
        header_rows: list[tuple[int, dict[str, int]]] = []
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            headers = {
                str(val).strip().lower(): pos
                for pos, val in enumerate(row)
                if val is not None and str(val).strip()
            }
            if {"datum", "belopp"}.issubset(headers.keys()):
                header_rows.append((idx, headers))
        if not header_rows:
            return 0, Decimal("0")
        header_idx, headers = header_rows[-1]
        date_idx = headers["datum"]
        amount_idx = headers["belopp"]
    else:
        raise AssertionError(f"Unexpected bank_import_type: {bank_import_type}")

    count = 0
    total = Decimal("0")
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if idx <= header_idx:
            continue
        date_value = row[date_idx] if date_idx < len(row) else None
        iso_date = _coerce_iso_date(date_value)
        if iso_date is None:
            continue
        amount_value = row[amount_idx] if amount_idx < len(row) else None
        if amount_value in (None, ""):
            continue
        total += Decimal(str(amount_value))
        count += 1
    return count, total


@pytest.mark.parametrize(
    ("bank_import_type", "filename"),
    [
        ("swedbank", "swedbank transactions nov 2025.xlsx"),
        ("seb", "seb tranasctions nov 2025.xlsx"),
        ("circle_k_mastercard", "circle k transactions nov 2025.xlsx"),
    ],
)
def test_preview_parses_sample_bank_statement(bank_import_type: str, filename: str):
    fixtures_dir = Path(__file__).resolve().parent / "fixtures" / "bank-transactions"
    file_path = fixtures_dir / filename
    assert file_path.exists(), f"Missing sample file: {file_path}"

    account_id = _create_account(bank_import_type=bank_import_type)
    payload = _b64_file(file_path)

    response = preview_imports(
        {
            "body": json.dumps(
                {
                    "files": [
                        {
                            "filename": filename,
                            "account_id": str(account_id),
                            "content_base64": payload,
                        }
                    ]
                }
            ),
            "isBase64Encoded": False,
        },
        None,
    )
    assert response["statusCode"] == 200
    body = _json_body(response)
    assert body["files"][0]["bank_import_type"] == bank_import_type
    assert body["files"][0]["error_count"] == 0
    assert body["files"][0]["row_count"] > 0
    assert len(body["rows"]) == body["files"][0]["row_count"]

    expected_count, expected_sum = _expected_transaction_count_and_sum(
        bank_import_type=bank_import_type,
        file_path=file_path,
    )
    assert expected_count == len(body["rows"])

    parsed_sum = sum((Decimal(row["amount"]) for row in body["rows"]), start=Decimal("0"))
    if bank_import_type == "circle_k_mastercard":
        assert parsed_sum == -expected_sum
        assert all(Decimal(row["amount"]) <= 0 for row in body["rows"])
    else:
        assert parsed_sum == expected_sum

    first = body["rows"][0]
    assert first["account_id"] == str(account_id)
    assert first["occurred_at"]
    assert first["description"]
    Decimal(first["amount"])
